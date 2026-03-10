from flask import Blueprint, request, jsonify, send_file
from io import BytesIO
from db import db
from routes.auth import admin_required
import openpyxl

admin_bp = Blueprint("admin", __name__)

PER_PAGE = 50


@admin_bp.route("/api/admin/logs")
@admin_required
def get_logs():
    try:
        page = int(request.args.get("page", 1))
    except (ValueError, TypeError):
        page = 1
    action = request.args.get("action", "")
    entity = request.args.get("entity", "")

    query = {}
    if action:
        query["action"] = action
    if entity:
        query["entity"] = entity

    total = db.activity_logs.count_documents(query)
    pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    logs = list(
        db.activity_logs.find(query, {"_id": 0})
        .sort("id", -1)
        .skip((page - 1) * PER_PAGE)
        .limit(PER_PAGE)
    )
    return jsonify({"logs": logs, "total": total, "pages": pages, "page": page})


@admin_bp.route("/api/admin/export/data")
@admin_required
def export_data():
    wb = openpyxl.Workbook()

    # Projects sheet
    ws = wb.active
    ws.title = "Projects"
    ws.append(["ID", "Client Name", "Project Name", "Total Amount", "Email", "Phone", "Notes"])
    for p in db.projects.find({}, {"_id": 0}).sort("id", 1):
        ws.append([p.get("id"), p.get("client_name"), p.get("project_name"),
                    p.get("total_amount"), p.get("email"), p.get("phone"), p.get("notes")])

    # Payments sheet
    ws2 = wb.create_sheet("Payments")
    ws2.append(["ID", "Project ID", "Date", "Paid Amount", "Note"])
    for pay in db.payments.find({}, {"_id": 0}).sort("id", 1):
        ws2.append([pay.get("id"), pay.get("project_id"), pay.get("date"),
                     pay.get("paid_amount"), pay.get("note")])

    # Employees sheet
    ws3 = wb.create_sheet("Employees")
    ws3.append(["ID", "Name", "Employee Code", "Department", "Salary/Month"])
    for emp in db.employees.find({}, {"_id": 0}).sort("id", 1):
        ws3.append([emp.get("id"), emp.get("name"), emp.get("employee_code"),
                     emp.get("department", ""), emp.get("salary_per_month")])

    # Salary Payments sheet
    ws4 = wb.create_sheet("Salary Payments")
    ws4.append(["ID", "Employee ID", "Month", "Date", "From Date", "To Date", "Amount Paid", "Status", "Note"])
    for sp in db.salary_payments.find({}, {"_id": 0}).sort("id", 1):
        ws4.append([sp.get("id"), sp.get("employee_id"), sp.get("month"),
                     sp.get("date"), sp.get("from_date", ""), sp.get("to_date", ""),
                     sp.get("amount_paid"), sp.get("status"), sp.get("note")])

    # Expenses sheet
    ws5 = wb.create_sheet("Expenses")
    ws5.append(["ID", "Date", "Description", "Category", "Amount", "Notes"])
    for exp in db.expenses.find({}, {"_id": 0}).sort("id", 1):
        ws5.append([exp.get("id"), exp.get("date"), exp.get("description"),
                     exp.get("category"), exp.get("amount"), exp.get("notes")])

    # Reminders sheet
    ws6 = wb.create_sheet("Reminders")
    ws6.append(["ID", "Description", "Date", "Time", "Email Sent"])
    for rem in db.reminders.find({}, {"_id": 0}).sort("id", 1):
        ws6.append([rem.get("id"), rem.get("description"), rem.get("date"),
                     rem.get("time"), rem.get("email_sent")])

    # Documents sheet
    ws7 = wb.create_sheet("Documents")
    ws7.append(["ID", "Client", "Project", "Date", "Type", "Status", "Created At"])
    for doc in db.documents.find({}, {"_id": 0}).sort("id", 1):
        ws7.append([doc.get("id"), doc.get("client_name"), doc.get("project_name"),
                     doc.get("date"), doc.get("doc_type"), doc.get("status"), doc.get("created_at")])

    # Document Actions sheet
    ws8 = wb.create_sheet("Document Actions")
    ws8.append(["ID", "Document ID", "Action", "Date", "User"])
    for da in db.document_actions.find({}, {"_id": 0}).sort("id", 1):
        ws8.append([da.get("id"), da.get("document_id"), da.get("action"),
                     da.get("date"), da.get("user")])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="technoji_business_data.xlsx")


@admin_bp.route("/api/admin/export/logs")
@admin_required
def export_logs():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activity Logs"
    ws.append(["ID", "Timestamp", "Username", "Action", "Entity", "Entity ID", "Details", "IP Address"])
    for log in db.activity_logs.find({}, {"_id": 0}).sort("id", -1):
        ws.append([log.get("id"), log.get("timestamp"), log.get("username"),
                    log.get("action"), log.get("entity"), log.get("entity_id"),
                    log.get("details"), log.get("ip_address")])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="technoji_activity_logs.xlsx")
